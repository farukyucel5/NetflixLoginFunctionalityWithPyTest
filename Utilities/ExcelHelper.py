import openpyxl


class ExcelClass:

    def excel_listeler_listesine_cevir(dosyayeri, pageName):
        filename = openpyxl.load_workbook(dosyayeri)
        page = filename[pageName]
        count_of_the_rows = page.max_row
        count_of_the_columns = page.max_column
        data = []

        for i in range(2, count_of_the_rows + 1):
            satir = []
            for j in range(1, count_of_the_columns + 1):
                if page.cell(i, j).value is None:
                    page.cell(i, j).value = ""

                satir.append(page.cell(i, j).value)
            data.append(satir)
        return data

